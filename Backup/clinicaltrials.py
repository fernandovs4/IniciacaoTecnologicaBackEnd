import requests
import openpyxl
from openpyxl import Workbook
from constantes import  *
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill
dicDeEstudos = {}
wb = Workbook()
ws = wb.active


ws.title = "Estudos Farmacéuticos"
ws["A1"] = "Pharmaceutical"
ws['A1'].font = Font(size=9, bold=True)
ws['A1'].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
for (i, letra) in enumerate(alfabeto_com_numero[1:]):
    ws[letra] = hosp[i]
    ws[letra].font = Font(size=9, bold=True)
    ws[letra].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

i = 2
import ast
with open('EstudoPorFarma.txt', 'r') as f:
    f = f.read() 
    f = ast.literal_eval(f)
    for farma, hospi in f.items():
            if not  ( hospi['A.C Camargo'] == 0 and hospi['Sírio Libanês'] == 0 and hospi['Albert Einstein'] == 0 and hospi['Moinhos de Vento'] == 0 and hospi['Beneficencia Portuguesa'] == 0 and hospi['Instituto Nacional do Cancer'] == 0 and hospi['CEPHO '] == 0 and hospi['Instituto Brasileiro de Controle do Cancer'] == 0 and hospi['PUCRS'] == 0 and hospi['CEPON'] == 0 and hospi['Hospital do Cancer de Barretos'] == 0 and hospi['Faculdade de Medicina do ABC'] == 0 and hospi['FMUSP'] == 0 and hospi['COI '] == 0 and hospi['ICESP'] == 0 and hospi['Alexander Fleming'] == 0 and hospi["D'or"] == 0 and hospi['Erasto Gaertner'] == 0 and hospi['Araujo Jorge'] == 0 and hospi['Senhora da Conceica'] == 0 and hospi['Mae de Deus'] == 0 and hospi['Oswaldo Cruz'] == 0 and hospi['Caridade de Ijui'] == 0 ):
        
                ws['A' + str(i)] = farma
                for (j, letra) in enumerate(alfabeto[1:]):
                    ws[letra + str(i)] = hospi[hosp[j]]
                  
    
                i += 1


with open("FarmaPorCondicao.txt", "r") as f:
    f = f.read()
    f = ast.literal_eval(f)
    ws3 = wb.create_sheet("Estudo por Condicao")
    ws3['A1'] = 'Condicao'
    ws3['A1'].font = Font(size=9, bold=True)
    ws3['A1'].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    for (i, letra) in enumerate(alfabeto_com_numero[1:]):
        ws3[letra] = hosp[i]
        ws3[letra].font = Font(size=9, bold=True)
        ws3[letra].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    i = 2
    for condicao, hospitais in f.items():
        ws3['A' + str(i)] = condicao
        for (j, letra) in enumerate(alfabeto[1:]):
            ws3[letra + str(i)] = hospitais[hosp[j]]
        i += 1

                

with open("EstudoPorData.txt", "r") as f:
    f = f.read() 
    f = ast.literal_eval(f)
   

ws = wb.create_sheet("Estudos por ano")
ws2 = wb.create_sheet("DASH - Por Captação")
ws['A1'] = 'Ano'
ws['A1'].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
ws['A1'].font = Font(size=9, bold=True)
ws['B1'] = 'TOTAL'
ws['B1'].font = Font(size=9, bold=True)
ws['B1'].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
for (i, letra) in enumerate(alfabeto_com_numero[2:]):
        ws[letra] = hosp[i]
        ws[letra].font = Font(size=9, bold=True)
        ws[letra].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        ws[letra].alignment = Alignment(wrap_text=True)

ws2['A1'] = 'Ano'
ws2['A1'].font = Font(size=9, bold=True)
ws2['A1'].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
ws2['B1'] = 'TOTAL'
ws2['B1'].font = Font(size=9, bold=True)
ws2['B1'].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
for (i, letra) in enumerate(alfabeto_com_numero[2:]):
        ws2[letra] = hosp[i]
        ws2[letra].font = Font(size=9, bold=True)
        ws2[letra].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        ws2[letra].alignment = Alignment(wrap_text=True)


i = 2
for  ano, estudo in f.items():
    ws['A' + str(i)] = ano
    ws['B' + str(i)] = sum(estudo.values())
    for letra in alfabeto[2:]:
        ws[letra + str(i)] = estudo[hosp[alfabeto.index(letra)-2]]
    
    ws2['A' + str(i)] = ano
    for letra in alfabeto[1:]:
        ws2[letra + str(i)] = estudo[hosp[alfabeto.index(letra)-1]]/sum(estudo.values())
        ws2[letra + str(i)].number_format = '0.000%'
        if letra == "C":
             ws2[letra + str(i)].font = Font(color= 'FF0000')
    i += 1
i = len(alfabeto)
for letra in alfabeto[1:]:
     ws[letra + str(i)] = '=SUM(' + letra + '2:' + letra + str(i-1) + ')' 
ws['A'  + str(i)] = 'TOTAL'
wb.save('arquivo5.xlsx')
   
    

















